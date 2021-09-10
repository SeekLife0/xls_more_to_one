#coding=utf-8
import sys
import xlrd
import os
import openpyxl
import time
import pythoncom
import win32com.client as win32

reload(sys)
sys.setdefaultencoding('utf-8')


# 读取路径下的所有文件名
def read_file_name(file_dir):
    L = []
    for root,dirs,files in os.walk(file_dir):
        for file in files:
            L.append(file)
    return L

# 进行xls到xlsx的转换
def xlstoxlsx(filePath):
    if filePath.find('xls'):
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        filePath1 = filePath.decode('utf-8').encode('gbk')
        wb = None
        try:
            wb = excel.Workbooks.Open(filePath1)
            wb.SaveAs(filePath1 + 'x', FileFormat=51)  # FileFormat = 51 is for .xlsx extension
            print "xls转为xlsx的绝对路径为"+filePath + 'x'
        except IOError:
            print '文件读写错误'
        finally:
            wb.Close()
            excel.Application.Quit()
        time.sleep(2)  # 避免未及时关闭的情况,等待关闭完成
        return filePath1 + 'x'

# 进行xlsx到xls的转换
def xlsxtoxls(file,filePath):
    if file.find('xlsx'):
        pythoncom.CoInitialize()
        excel = win32.gencache.EnsureDispatch('Excel.Application')
        file = file.decode('utf-8').encode('gbk')
        # 去除末尾的x[-1:]
        # filepaths = filePath.split('.')
        # filePath1 = filepaths[0]
        filePath1 = "{}".format(filePath[:-4])
        filePath1 = filePath1.decode('utf-8').encode('gbk')
        # 打开文件需要进行异常捕获和关闭文件
        wb = None
        try:
            wb = excel.Workbooks.Open(file)
            wb.SaveAs(filePath1 + '.xls', FileFormat=56)  # FileFormat = 51 is for .xlsx extension
            print filePath1 + '.xls'
        except IOError:
            print '该文件读写发生错误'
        finally:
            wb.Close()
            excel.Application.Quit()
        time.sleep(1)  # 避免未及时关闭的情况,等待关闭完成
        return filePath1 + '.xls'

# 写入到execel表格
# 先打开原表然后复制一份最后另存为
# values为单表内容总和 startCol为开始的列数 row为开始的行数 col为爬取的内容数
def wExecel(path,row,col,startCol,values):
    # 这里最开始先判断一下填的路径是否存在
    print "是否存在路径判断"+path
    if os.path.isfile(path.decode('utf-8').encode('gbk')):
        # 这里需要对路径文件格式进行判断一个if语句
        if path.find('xlsx') != -1:                    # 把xlsx转为xls，会把原路径的xlsx文件替换
            xlsxPath = path.decode('utf-8').encode('gbk')
            print xlsxPath.decode('gbk').encode('utf-8')   #为了方便控制台查看转为utf-8
        else:
            print '文件格式为：xls'
            xlsxPath = xlstoxlsx(path)
            # 转换完成之后需要把原来xls这个文件删除
            path = path.decode('utf-8').encode('gbk')
            # os.remove(path)
            # print '原xls文件已删除'
        # wb = None
        try:
            wb = openpyxl.load_workbook(xlsxPath) # 打开这个xlsx文件得到wb对象
            sheetnames = wb.get_sheet_names()     # 获取文件内的所有表格
            # sheet = wb.get_sheet_by_name(sheetnames[0])
            for sheetname in sheetnames:          # 遍历这个文件的内的所有表格 这里只取第一个表
                sheet = wb.get_sheet_by_name(sheetname)       # 取sheet表
                # print "遍历的表格数" + str(j)
                # range(start,stop,step) 开始 结束（不包括） 步长
                for num in range(0, len(values)-(col-1), col):   # col每次遍历跳过的数 以col的数量为单位 len(values)一定是col的整数倍
                    for a in range(0, col):                      # 对一行数据进行操作 一共col列循环col次
                        value = str(values[num+a])               # 获取内容
                        # value = value.decode('utf-8').encode('gbk')
                        print "写入的内容" + value
                        sheet.cell(row, startCol+a).value = value  # 把内容写入到汇总表对应位置
                    row += 1
                break
                # j += 1
            # 这里保存的路径需要修改，只需要路径名
            print "最终保存路径---->"+xlsxPath.decode('gbk').encode('utf-8')       #为了方便控制台查看转为utf-8
            wb.save(xlsxPath)
        except IOError:
            print 'IO读写错误----->107'
    else:
        print "你填入的导出路径不存在请重新填写"

# 该函数抓取导入的多张表的某些指定内容，内容指定由坐标表示
# 因为涉及到一个文件可能存在多个表的问题所以要以表为单位进行内容抓取
def rwExecel(Rpath,eList_col,eList_entry):
    # global list_col
    xlsPath = Rpath
    xls = None
    # 这里需要对路径文件格式进行判断一个if语句
    if 'xlsx' in Rpath:                    # 把xlsx转为xls，会把原路径的xlsx文件替换为xls,读的时候只能操作xls表格
        print '文件格式为：xlsx'
        xlsPath = xlsxtoxls(Rpath,Rpath)
        # 转换完成之后需要把原来xlsx这个文件删除
        path = Rpath.decode('utf-8').encode('gbk')
        os.remove(path)
        print '文件已删除'
        print xlsPath.decode('gbk').encode('utf-8')       #为了方便控制台查看转为utf-8
        # xls = xlrd.open_workbook(xlsPath)
    else:
        # xlsPath = xlsPath.decode('utf-8').encode('gbk')
        print xlsPath.decode('gbk').encode('utf-8')       #为了方便控制台查看转为utf-8
    # --------对xlsx文件内的所有表格都进行判断是否有内容-----------
    xls = xlrd.open_workbook(xlsPath)
    tables = xls.sheets()         # 获取文件所有表格这样就不需要表的标号和名称了
    sheet_name = xls.sheet_names()
    print sheet_name
    for table in tables:         # 如何判断一个表是否有内容
        if table.nrows == 0:     # 查看表的有效函数是否为0
            continue             # 直接进入下一个循环
        # 通过遍历elist_entry来获得对应坐标,每三个为一个单位获取
        for num in range(0,len(eList_entry)-2,3):
            # 只读入有效坐标只有行或列或者全没有不读入直接跳入下一个循环
            if (eList_entry[num+1]!='' and eList_entry[num+2]!='') and (eList_entry[num+1]!=None and eList_entry[num+2]!=None):
                r = int(eList_entry[num+1])   # 第二个是行
                c = int(eList_entry[num+2])   # 第三个是列
                eList_col.append(table.cell_value(r,c)) # 爬取表格内容
                print '每次爬取内容' + str(table.cell_value(r,c))
            else:
                continue
        # wExecel(Wpath,eRow,eCol,eList_col) #写入的文件路径
        # del eList_col[:]  #每次执行完一行清空列表
        # eRow = eRow + 1
    return eList_col  # 返回装载所有多表信息的列表

# 爬取表格内容并复制到表格力功能A
# 多个表格内容复制到另一个表，一个表的内容就是一行
def deal_Excel_A(start_row_col,content_row_col,importPath,exportPath):
    print u"开始执行任务"
    print '用户输入行数'+start_row_col[0]
    row = int(start_row_col[0]) + 1                       # 获得用户输入的开始行数
    print '用户输入列数'+start_row_col[1]
    col = int(start_row_col[1]) + 1                   # 获得用户输入的开始列数
    list_col = []                                     # 用来保存一个人的表格内容
    list_entry = content_row_col                      # 获得用户输入的行和列
    allFilesNames = read_file_name(importPath)        # 获取每个xls的表名
    i = 0
    for fileName in allFilesNames:                    # 遍历每个表名
        print fileName.decode('gbk').encode('utf-8')                               # 打印获得的每个表名 #为了方便控制台查看转为utf-8
        list_col = rwExecel(importPath + '/' + fileName,
                            list_col, list_entry)     # 抓取一张表的内容然后暂存到list_col
        i += 1
    print '遍历的表格文件数目---->' + str(i)
    # 读内容到列表完毕后进行统一写入操作
    wExecel(exportPath, row, len(list_entry)/3, col, list_col) # 把list_col的内容全部写入新的xlsx文件内

# -------需要提供的数据--------
start_row_col = ["1","1"]          # 最终导入的单表->数组第0个为开始行 数组第1个为开始列
content_row_col = ["姓名","0","1", "年龄","0","3", "出生日期","1","1"]        # 抓取的坐标以及内容名称（表明这个坐标抓取的东西方便查看）



if __name__ == '__main__':
    deal_Excel_A(start_row_col,content_row_col,"I:/pythonProject/xls_operation/importPath",
                 "I:/pythonProject/xls_operation/exportPath/某机构汇总表.xls")